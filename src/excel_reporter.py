"""
AI Attendance Trend Analyzer - Automated Excel Reporter
Formats student performance and trend analytics into Excel workbooks with conditional styling.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_excel_report(df_raw, df_students, kpis):
    output = io.BytesIO()
    wb = Workbook()

    # Styling Palette
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Red
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Yellow
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Sheet 1: Student Risk & Trend Summary
    ws_summary = wb.active
    ws_summary.title = "Attendance Risk Analysis"
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary["A1"] = "AI-POWERED STUDENT ATTENDANCE TREND & RISK REPORT"
    ws_summary["A1"].font = title_font

    # KPI Block
    kpi_items = [
        ("Total Students", kpis.get("Total_Students", 0)),
        ("Avg Attendance", f"{kpis.get('Avg_Attendance_Pct', 0)}%"),
        ("High Risk (<75%)", kpis.get("High_Risk_Count", 0)),
        ("Warning / Declining", kpis.get("Warning_Count", 0)),
        ("Safe (>80%)", kpis.get("Safe_Count", 0)),
        ("Attendance-GPA Correlation", kpis.get("Attendance_GPA_Correlation", 0))
    ]

    for col_idx, (lbl, val) in enumerate(kpi_items, start=1):
        cell_lbl = ws_summary.cell(row=3, column=col_idx, value=lbl)
        cell_lbl.font = Font(name="Calibri", size=9, bold=True, color="595959")
        cell_lbl.alignment = Alignment(horizontal="center")

        cell_val = ws_summary.cell(row=4, column=col_idx, value=val)
        cell_val.font = Font(name="Calibri", size=13, bold=True)
        cell_val.alignment = Alignment(horizontal="center")
        cell_val.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell_val.border = thin_border

    # Student Table
    start_row = 7
    for r_idx, row in enumerate(dataframe_to_rows(df_students, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border

            if r_idx == start_row:
                cell.fill = navy_fill
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = Font(name="Calibri", size=10)
                # Conditional formatting for Risk Status column
                if c_idx == list(df_students.columns).index("Risk_Level") + 1:
                    if val == "HIGH":
                        cell.fill = red_fill
                        cell.font = Font(name="Calibri", size=10, bold=True, color="C00000")
                    elif val == "MEDIUM":
                        cell.fill = yellow_fill
                        cell.font = Font(name="Calibri", size=10, bold=True, color="B25900")
                    elif val == "LOW":
                        cell.fill = green_fill
                        cell.font = Font(name="Calibri", size=10, color="375623")

    # Sheet 2: Raw Monthly Attendance Records
    ws_raw = wb.create_sheet(title="Raw Monthly Logs")
    ws_raw.views.sheetView[0].showGridLines = True
    ws_raw["A1"] = "MULTI-SEMESTER MONTHLY ATTENDANCE DATA"
    ws_raw["A1"].font = title_font

    start_raw = 3
    for r_idx, row in enumerate(dataframe_to_rows(df_raw, index=False, header=True), start=start_raw):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_raw.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if r_idx == start_raw:
                cell.fill = navy_fill
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = Font(name="Calibri", size=10)

    # Column Widths
    for ws in [ws_summary, ws_raw]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output)
    output.seek(0)
    return output.getvalue()
